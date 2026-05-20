#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Captures a leading variable name, then a colon/en-dash/hyphen separator, then the rest as question text.
COMBINED_HEADER = re.compile(r"^([A-Za-z][A-Za-z0-9_.-]*)\s*[:–-]\s*(.+)\s*$")

# Captures a Decipher-style variable like Q1A1_1 followed by whitespace and a longer question stem.
DECIPHER_STYLE = re.compile(r"^(Q\d+[A-Za-z0-9_.-]*)\s+(.{10,})\s*$")

# Treats a token as a bare variable identifier when it matches this pattern.
BARE_ID = re.compile(r"^[A-Za-z][A-Za-z0-9_.-]*$")


def is_probably_question_row(values: list[object]) -> bool:
    strs = [str(v).strip() for v in values if v not in (None, "")]
    if len(strs) < 3:
        return False
    longish = sum(1 for s in strs if len(s) > 60)
    return longish >= max(3, len(strs) // 4)


def stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def split_header_cell(cell: str) -> tuple[str, str]:
    if not cell:
        return "", ""
    m = COMBINED_HEADER.match(cell)
    if m:
        return m.group(1), m.group(2).strip()
    m = DECIPHER_STYLE.match(cell)
    if m:
        return m.group(1), m.group(2).strip()
    if BARE_ID.match(cell):
        return cell, ""
    return "", cell


def normalize_headers(row1: list[str], row2: list[str], row2_is_questions: bool) -> tuple[list[str], list[str]]:
    padded_len = max(len(row1), len(row2))
    row1 = list(row1) + [""] * (padded_len - len(row1))
    row2 = list(row2) + [""] * (padded_len - len(row2))

    ids: list[str] = []
    texts: list[str] = []
    auto_i = 1

    for top, second in zip(row1, row2, strict=True):
        id_from_top, q_from_top = split_header_cell(top)

        if id_from_top and q_from_top:
            ids.append(id_from_top)
            texts.append(q_from_top)
            continue

        if id_from_top and not q_from_top:
            ids.append(id_from_top)
            texts.append(second if row2_is_questions and second else "")
            continue

        if not id_from_top and top:
            ids.append(f"AUTO_Q{auto_i:03d}")
            auto_i += 1
            texts.append(top)
            continue

        if second and not top:
            if row2_is_questions:
                ids.append(f"AUTO_Q{auto_i:03d}")
                auto_i += 1
                texts.append(second)
            else:
                ids.append("")
                texts.append("")
            continue

        ids.append("")
        texts.append("")

    return ids, texts


def read_sheet_rows(ws: Worksheet, max_col: int, max_row: int) -> list[list[object]]:
    rows: list[list[object]] = []
    for r in range(1, max_row + 1):
        row: list[object] = []
        for c in range(1, max_col + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)
    return rows


def write_row(ws: Worksheet, row_idx: int, values: list[str]) -> None:
    for c, v in enumerate(values, start=1):
        ws.cell(row_idx, c).value = v if v != "" else None


def transform_workbook(path_in: Path, path_out: Path) -> None:
    wb = load_workbook(path_in)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        raise ValueError("Worksheet is empty")

    rows = read_sheet_rows(ws, max_col=max_col, max_row=max_row)
    row1 = [stringify(v) for v in rows[0]]
    row2 = [stringify(v) for v in rows[1]] if len(rows) >= 2 else []

    row2_is_questions = is_probably_question_row(rows[1]) if len(rows) >= 2 else False

    ids, texts = normalize_headers(row1, row2, row2_is_questions=row2_is_questions)

    if row2_is_questions:
        data_start_old = 3
    else:
        data_start_old = 2

    out = Workbook()
    ows = out.active
    write_row(ows, 1, ids)
    write_row(ows, 2, texts)

    out_row = 3
    for old_r in range(data_start_old, max_row + 1):
        for c in range(1, max_col + 1):
            ows.cell(out_row, c).value = ws.cell(old_r, c).value
        out_row += 1

    out.save(path_out)


def write_xlsx_from_csv(path_in: Path, path_out: Path) -> None:
    text = path_in.read_text(encoding="utf-8", errors="replace").splitlines()
    if len(text) < 3:
        raise ValueError("CSV must contain a header row, a question row, and at least one respondent row")

    sample = text[0]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","

    rows: list[list[str]] = []
    with path_in.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            rows.append(row)

    max_col = max(len(r) for r in rows)
    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(rows, start=1):
        padded = list(row) + [""] * (max_col - len(row))
        for c_idx, v in enumerate(padded, start=1):
            ws.cell(r_idx, c_idx).value = v if v != "" else None

    wb.save(path_out)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Normalize a ReDem import sheet to Row 1 = variable IDs and Row 2 = question texts.",
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--from-csv",
        action="store_true",
        help="Treat input as CSV with an ID row, a question row, then respondent rows (no header splitting).",
    )
    args = parser.parse_args()

    if args.from_csv:
        write_xlsx_from_csv(args.input, args.output)
    else:
        transform_workbook(args.input, args.output)


if __name__ == "__main__":
    main()
